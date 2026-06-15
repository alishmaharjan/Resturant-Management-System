import csv
import json
from decimal import Decimal
from datetime import timedelta
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Prefetch
from django.utils import timezone
from apps.orders.models import Order, OrderItem
from apps.billing.models import Payment, Refund, CreditRecord
from apps.reports.models import AuditLog


def _date_range(request, default_range='today'):
    """Parse start/end datetimes from request GET params or range preset."""
    today = timezone.now().date()
    tz    = timezone.get_current_timezone()

    range_opt  = request.GET.get('range', default_range)
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    if date_from and date_to:
        try:
            from datetime import date
            date_from = date.fromisoformat(date_from)
            date_to   = date.fromisoformat(date_to)
            range_opt = 'custom'
        except ValueError:
            date_from = date_to = today
            range_opt = default_range
    elif range_opt == 'week':
        date_from = today - timedelta(days=6)
        date_to   = today
    elif range_opt == 'month':
        date_from = today.replace(day=1)
        date_to   = today
    else:
        date_from = date_to = today

    start = timezone.datetime.combine(date_from, timezone.datetime.min.time()).replace(tzinfo=tz)
    end   = timezone.datetime.combine(date_to,   timezone.datetime.max.time()).replace(tzinfo=tz)
    return date_from, date_to, start, end, range_opt


@login_required(login_url='/login/')
def reports_index(request):
    date_from, date_to, start, end, range_opt = _date_range(request)

    paid_orders = Order.objects.filter(status='PAID', created_at__gte=start, created_at__lte=end)
    agg         = paid_orders.aggregate(revenue=Sum('grand_total'), count=Count('id'))
    revenue     = agg['revenue'] or 0
    order_count = agg['count']   or 0
    avg_order   = (revenue / order_count) if order_count else 0

    refund_agg    = Refund.objects.filter(refunded_at__gte=start, refunded_at__lte=end).aggregate(
        total=Sum('amount'), count=Count('id')
    )
    total_refunds = refund_agg['total'] or 0
    refund_count  = refund_agg['count'] or 0
    net_revenue   = revenue - total_refunds

    top_items = (
        OrderItem.objects
        .filter(order__status='PAID', order__created_at__gte=start)
        .values('product__name', 'product__category__name')
        .annotate(qty=Sum('qty'), revenue=Sum('line_total'))
        .order_by('-revenue')[:15]
    )

    pay_breakdown = (
        Payment.objects
        .filter(paid_at__gte=start, paid_at__lte=end)
        .values('method')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    # Last 7 days for chart
    chart_labels, chart_data = [], []
    for i in range(6, -1, -1):
        d   = date_to - timedelta(days=i)
        tz  = timezone.get_current_timezone()
        ds  = timezone.datetime.combine(d, timezone.datetime.min.time()).replace(tzinfo=tz)
        de  = timezone.datetime.combine(d, timezone.datetime.max.time()).replace(tzinfo=tz)
        rev = Order.objects.filter(
            status='PAID', created_at__gte=ds, created_at__lte=de
        ).aggregate(t=Sum('grand_total'))['t'] or 0
        chart_labels.append(d.strftime('%d %b'))
        chart_data.append(float(rev))

    recent_refunds = (
        Refund.objects
        .filter(refunded_at__gte=start, refunded_at__lte=end)
        .select_related('payment__order')
        .order_by('-refunded_at')[:50]
    )

    audit_logs = AuditLog.objects.all()[:50]

    return render(request, 'reports/index.html', {
        'range_opt':      range_opt,
        'date_from':      date_from,
        'date_to':        date_to,
        'revenue':        revenue,
        'net_revenue':    net_revenue,
        'total_refunds':  total_refunds,
        'refund_count':   refund_count,
        'order_count':    order_count,
        'avg_order':      avg_order,
        'top_items':      top_items,
        'pay_breakdown':  pay_breakdown,
        'recent_refunds': recent_refunds,
        'chart_labels':  json.dumps(chart_labels),
        'chart_data':    json.dumps(chart_data),
        'audit_logs':    audit_logs,
    })


# ── CSV Exports ────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def export_sales_csv(request):
    """Download sales report: one row per paid order."""
    date_from, date_to, start, end, _ = _date_range(request)
    orders = (
        Order.objects
        .filter(status='PAID', created_at__gte=start, created_at__lte=end)
        .prefetch_related('payments__refunds', 'credit_records__account', 'items')
        .annotate(item_count=Count('items'))
        .order_by('created_at')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="yasumi_sales_{date_from}_{date_to}.csv"'
    )
    w = csv.writer(response)
    w.writerow(['Order No', 'Date', 'Time', 'Type', 'Table', 'Items',
                'Subtotal', 'Discount', 'Grand Total', 'Payment Methods',
                'Refunded (Rs.)', 'Net Collected'])
    for o in orders:
        methods = []
        for p in o.payments.all():
            methods.append(p.method)
        for cr in o.credit_records.all():
            methods.append(f'CREDIT ({cr.account.name})')
        refunded = sum(
            r.amount for p in o.payments.all() for r in p.refunds.all()
        )
        w.writerow([
            o.order_no,
            o.created_at.strftime('%Y-%m-%d'),
            o.created_at.strftime('%H:%M'),
            o.order_type,
            o.table_no or '—',
            o.item_count,
            o.subtotal,
            o.discount_amount,
            o.grand_total,
            ', '.join(methods) or '—',
            refunded or '—',
            o.grand_total - refunded,
        ])
    return response


@login_required(login_url='/login/')
def export_order_items_csv(request):
    """Download detailed order items: one row per item in each paid order."""
    date_from, date_to, start, end, _ = _date_range(request)
    items = (
        OrderItem.objects
        .filter(order__status='PAID', order__created_at__gte=start, order__created_at__lte=end)
        .select_related('order', 'product', 'product__category')
        .order_by('order__created_at', 'id')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="yasumi_order_items_{date_from}_{date_to}.csv"'
    )
    w = csv.writer(response)
    w.writerow(['Order No', 'Date', 'Type', 'Table',
                'Category', 'Item', 'Qty', 'Unit Price', 'Line Total',
                'Order Subtotal', 'Discount', 'Order Total'])
    for i in items:
        o = i.order
        w.writerow([
            o.order_no,
            o.created_at.strftime('%Y-%m-%d %H:%M'),
            o.order_type,
            o.table_no or '—',
            i.product.category.name if i.product.category else '—',
            i.product.name,
            i.qty,
            i.unit_price,
            i.line_total,
            o.subtotal,
            o.discount_amount,
            o.grand_total,
        ])
    return response


@login_required(login_url='/login/')
def export_products_csv(request):
    """Download top products: qty sold and revenue per item."""
    date_from, date_to, start, end, _ = _date_range(request)
    rows = (
        OrderItem.objects
        .filter(order__status='PAID', order__created_at__gte=start, order__created_at__lte=end)
        .values('product__name', 'product__category__name')
        .annotate(qty_sold=Sum('qty'), revenue=Sum('line_total'))
        .order_by('-revenue')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="yasumi_products_{date_from}_{date_to}.csv"'
    )
    w = csv.writer(response)
    w.writerow(['Category', 'Item Name', 'Qty Sold', 'Revenue (Rs.)'])
    total_qty = total_rev = 0
    for r in rows:
        w.writerow([
            r['product__category__name'] or '—',
            r['product__name'],
            r['qty_sold'],
            r['revenue'],
        ])
        total_qty += r['qty_sold'] or 0
        total_rev += r['revenue'] or 0
    w.writerow([])
    w.writerow(['TOTAL', '', total_qty, total_rev])
    return response


@login_required(login_url='/login/')
def export_payments_csv(request):
    """Download payment transactions."""
    date_from, date_to, start, end, _ = _date_range(request)
    payments = (
        Payment.objects
        .filter(paid_at__gte=start, paid_at__lte=end)
        .select_related('order')
        .order_by('paid_at')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="yasumi_payments_{date_from}_{date_to}.csv"'
    )
    w = csv.writer(response)
    w.writerow(['Order No', 'Method', 'Amount (Rs.)', 'Reference', 'Date', 'Time'])
    total = 0
    for p in payments:
        w.writerow([
            p.order.order_no,
            p.method,
            p.amount,
            p.txn_ref or '—',
            p.paid_at.strftime('%Y-%m-%d'),
            p.paid_at.strftime('%H:%M'),
        ])
        total += p.amount
    w.writerow([])
    w.writerow(['TOTAL', '', total, '', '', ''])

    refunds = (
        Refund.objects
        .filter(refunded_at__gte=start, refunded_at__lte=end)
        .select_related('payment__order')
        .order_by('refunded_at')
    )
    if refunds.exists():
        w.writerow([])
        w.writerow(['--- REFUNDS ---', '', '', '', '', ''])
        w.writerow(['Order No', 'Reason', 'Refund Amount (Rs.)', '', 'Date', 'Time'])
        refund_total = 0
        for r in refunds:
            w.writerow([
                r.payment.order.order_no,
                r.reason or '—',
                r.amount,
                '',
                r.refunded_at.strftime('%Y-%m-%d'),
                r.refunded_at.strftime('%H:%M'),
            ])
            refund_total += r.amount
        w.writerow([])
        w.writerow(['REFUND TOTAL', '', refund_total, '', '', ''])
        w.writerow(['NET TOTAL', '', total - refund_total, '', '', ''])

    return response


# ── Master Report (Excel) ──────────────────────────────────────────────────────

@login_required(login_url='/login/')
def export_master_report(request):
    """Master Report — Excel workbook: Transactions sheet + Summary sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_from, date_to, start, end, _ = _date_range(request)

    orders = list(
        Order.objects
        .filter(status='PAID', created_at__gte=start, created_at__lte=end)
        .prefetch_related(
            Prefetch('payments', queryset=Payment.objects.prefetch_related('refunds')),
            Prefetch('credit_records', queryset=CreditRecord.objects.select_related('account')),
            Prefetch('items', queryset=OrderItem.objects.select_related('product')),
        )
        .order_by('created_at')
    )

    # ── Shared style constants ────────────────────────────────────────────────
    C_GREEN      = '2A5C3E'
    C_GREEN_DARK = '1A3D28'
    C_GREEN_PALE = 'EAF2EC'
    C_ALT        = 'F7FAF8'
    C_WHITE      = 'FFFFFF'

    H_FILL   = PatternFill('solid', fgColor=C_GREEN)
    TOT_FILL = PatternFill('solid', fgColor=C_GREEN_DARK)
    ALT_FILL = PatternFill('solid', fgColor=C_ALT)
    HI_FILL  = PatternFill('solid', fgColor=C_GREEN_PALE)

    def hdr_font(sz=9):
        return Font(name='Calibri', size=sz, bold=True, color=C_WHITE)

    def data_font(bold=False):
        return Font(name='Calibri', size=9, bold=bold)

    def tot_font():
        return Font(name='Calibri', size=9, bold=True, color=C_WHITE)

    MONEY    = '#,##0.00'
    CENTER   = Alignment(horizontal='center', vertical='center')
    LEFT     = Alignment(horizontal='left',   vertical='center')
    RIGHT    = Alignment(horizontal='right',  vertical='center')
    WRAP_L   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    thin_side = Side(style='thin', color='DDDDDD')
    ROW_BORDER = Border(bottom=thin_side)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    period_str = (
        f'{date_from.strftime("%d %b %Y")} to {date_to.strftime("%d %b %Y")}'
        if date_from != date_to else date_from.strftime('%d %b %Y')
    )
    generated = timezone.now().strftime('%d %b %Y  %H:%M')

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — TRANSACTIONS
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Transactions'
    ws.sheet_view.showGridLines = False

    # Title block (18 columns = A:R)
    ws.merge_cells('A1:R1')
    ws['A1'] = 'YASUMI RESTAURANT  ·  MASTER REPORT'
    ws['A1'].font      = Font(name='Calibri', size=15, bold=True, color=C_GREEN)
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:R2')
    ws['A2'] = f'Period: {period_str}     Generated: {generated}'
    ws['A2'].font      = Font(name='Calibri', size=9, italic=True, color='888888')
    ws['A2'].alignment = CENTER
    ws.row_dimensions[2].height = 15

    ws.row_dimensions[3].height = 6  # spacer

    # Credit-amount highlight fill (amber/gold tint so it stands out)
    CREDIT_FILL     = PatternFill('solid', fgColor='FFF8E1')
    CREDIT_FILL_ALT = PatternFill('solid', fgColor='FFF3CD')

    # Column definitions: (header, column width)
    # Col 14 = Credit Amount (Rs.) — new dedicated column
    TXN_COLS = [
        ('S.No',                  6),   # 1
        ('Date',                 13),   # 2
        ('Day',                   9),   # 3
        ('Time',                  7),   # 4
        ('Order No',             19),   # 5
        ('Type',                  9),   # 6
        ('Table',                15),   # 7
        ('Items',                 6),   # 8
        ('Items Detail',         42),   # 9
        ('Subtotal (Rs.)',       14),   # 10
        ('Discount (Rs.)',       14),   # 11
        ('Grand Total (Rs.)',    16),   # 12
        ('Payment Breakdown',    30),   # 13
        ('Credit Amount (Rs.)',  16),   # 14  ← new
        ('Credit Customer',      18),   # 15
        ('TXN Reference',        18),   # 16
        ('Refunded (Rs.)',       14),   # 17
        ('Net Collected (Rs.)',  16),   # 18
    ]
    N_COLS  = len(TXN_COLS)
    HDR_ROW = 4

    for c, (label, width) in enumerate(TXN_COLS, 1):
        cell = ws.cell(row=HDR_ROW, column=c, value=label)
        cell.font      = hdr_font()
        cell.fill      = H_FILL
        cell.alignment = CENTER
        cell.border    = ROW_BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
        # Give the credit column a gold header to draw the eye
        if c == 14:
            cell.fill = PatternFill('solid', fgColor='E65100')
    ws.row_dimensions[HDR_ROW].height = 22

    ws.freeze_panes = f'A{HDR_ROW + 1}'
    ws.auto_filter.ref = (
        f'A{HDR_ROW}:{get_column_letter(N_COLS)}{HDR_ROW}'
    )

    # Data rows
    tot_sub = tot_disc = tot_grand = tot_credit = tot_ref = tot_net = Decimal('0')
    # Columns that get money formatting
    MONEY_COLS  = {10, 11, 12, 14, 17, 18}
    CENTER_COLS = {1, 3, 6, 8}

    for sn, order in enumerate(orders, 1):
        dr   = HDR_ROW + sn
        fill = ALT_FILL if sn % 2 == 0 else None

        items_list   = list(order.items.all())
        items_detail = ', '.join(f'{i.product.name} ×{i.qty}' for i in items_list)

        # Build payment breakdown with amounts: "CASH: 1,200 | FONEPAY: 500"
        pay_parts    = []
        credit_amt   = Decimal('0')
        credit_name  = ''
        txn_ref      = ''
        for p in order.payments.all():
            pay_parts.append(f'{p.method}: {float(p.amount):,.2f}')
            if p.txn_ref:
                txn_ref = p.txn_ref
        for cr in order.credit_records.all():
            pay_parts.append(f'CREDIT: {float(cr.amount):,.2f}')
            credit_amt  += cr.amount
            credit_name  = cr.account.name

        refunded = sum(r.amount for p in order.payments.all() for r in p.refunds.all())
        net      = order.grand_total - refunded

        tot_sub    += order.subtotal
        tot_disc   += order.discount_amount
        tot_grand  += order.grand_total
        tot_credit += credit_amt
        tot_ref    += refunded
        tot_net    += net

        lt = timezone.localtime(order.created_at)
        row_vals = [
            sn,                                                      # 1
            lt.strftime('%d %b %Y'),                                 # 2
            lt.strftime('%a'),                                       # 3
            lt.strftime('%H:%M'),                                    # 4
            order.order_no,                                          # 5
            'Dine In' if order.order_type == 'DINE_IN' else 'Takeaway',  # 6
            order.table_no or '—',                                   # 7
            len(items_list),                                         # 8
            items_detail,                                            # 9
            float(order.subtotal),                                   # 10
            float(order.discount_amount),                            # 11
            float(order.grand_total),                                # 12
            ' | '.join(pay_parts) or '—',                           # 13
            float(credit_amt),                                       # 14  Credit Amount
            credit_name or '—',                                      # 15
            txn_ref or '—',                                          # 16
            float(refunded),                                         # 17
            float(net),                                              # 18
        ]

        for c, val in enumerate(row_vals, 1):
            cell        = ws.cell(row=dr, column=c, value=val)
            cell.font   = data_font()
            cell.border = ROW_BORDER

            # Credit Amount column: amber tint when > 0
            if c == 14:
                cell.fill          = CREDIT_FILL_ALT if sn % 2 == 0 else CREDIT_FILL
                cell.number_format = MONEY
                cell.alignment     = RIGHT
                if val and float(val) > 0:
                    cell.font = Font(name='Calibri', size=9, bold=True, color='E65100')
            elif fill:
                cell.fill = fill

            if c in MONEY_COLS and c != 14:
                cell.number_format = MONEY
                cell.alignment     = RIGHT
            elif c in CENTER_COLS:
                cell.alignment = CENTER
            elif c == 9:
                cell.alignment = WRAP_L
            elif c != 14:
                cell.alignment = LEFT
        ws.row_dimensions[dr].height = 18

    # Totals row
    tot_row = HDR_ROW + len(orders) + 1
    ws.merge_cells(f'A{tot_row}:I{tot_row}')
    tc = ws.cell(row=tot_row, column=1, value='TOTALS')
    tc.font = tot_font(); tc.fill = TOT_FILL; tc.alignment = CENTER

    for c, val in [
        (10, tot_sub), (11, tot_disc), (12, tot_grand),
        (14, tot_credit), (17, tot_ref), (18, tot_net),
    ]:
        cell = ws.cell(row=tot_row, column=c, value=float(val))
        cell.font          = tot_font()
        cell.fill          = TOT_FILL if c != 14 else PatternFill('solid', fgColor='BF360C')
        cell.number_format = MONEY
        cell.alignment     = RIGHT

    for c in [13, 15, 16]:
        ws.cell(row=tot_row, column=c).fill = TOT_FILL
    ws.row_dimensions[tot_row].height = 22

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 — SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 16

    # Title
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'YASUMI RESTAURANT  ·  SUMMARY'
    ws2['A1'].font      = Font(name='Calibri', size=15, bold=True, color=C_GREEN)
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells('A2:F2')
    ws2['A2'] = f'Period: {period_str}     Generated: {generated}'
    ws2['A2'].font      = Font(name='Calibri', size=9, italic=True, color='888888')
    ws2['A2'].alignment = CENTER
    ws2.row_dimensions[2].height = 15
    ws2.row_dimensions[3].height = 8

    r = 4

    def s2_section(title):
        nonlocal r
        ws2.row_dimensions[r].height = 8
        r += 1
        c = ws2.cell(row=r, column=1, value=title)
        c.font = Font(name='Calibri', size=11, bold=True, color=C_GREEN)
        ws2.row_dimensions[r].height = 20
        r += 1

    def s2_kv(label, value, highlight=False, is_count=False):
        nonlocal r
        lc = ws2.cell(row=r, column=1, value=label)
        vc = ws2.cell(row=r, column=2, value=(int(value) if is_count else float(value)))
        lc.font = Font(name='Calibri', size=10, bold=highlight)
        vc.font = Font(name='Calibri', size=10, bold=highlight)
        vc.alignment = RIGHT
        if not is_count:
            vc.number_format = MONEY
        if highlight:
            lc.fill = vc.fill = HI_FILL
        ws2.row_dimensions[r].height = 18
        r += 1

    def s2_hdr(*cols):
        nonlocal r
        for ci, label in enumerate(cols, 1):
            c = ws2.cell(row=r, column=ci, value=label)
            c.font = hdr_font(); c.fill = H_FILL; c.alignment = CENTER
        ws2.row_dimensions[r].height = 20
        r += 1

    def s2_row(*vals, alt=False):
        nonlocal r
        fill = ALT_FILL if alt else None
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=r, column=ci, value=val)
            c.font = data_font()
            if fill:
                c.fill = fill
            if isinstance(val, float):
                c.number_format = MONEY
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
        ws2.row_dimensions[r].height = 18
        r += 1

    order_count = len(orders)
    avg_order   = tot_grand / order_count if order_count else Decimal('0')

    # ── Financial overview ──────────────────────────────────────────────────
    s2_section('FINANCIAL OVERVIEW')
    s2_kv('Total Paid Orders',        order_count,  is_count=True)
    s2_kv('Gross Revenue (Rs.)',       tot_grand)
    s2_kv('Total Discounts Given (Rs.)', tot_disc)
    s2_kv('Total Refunds (Rs.)',       tot_ref)
    s2_kv('Net Revenue (Rs.)',         tot_net,  highlight=True)
    s2_kv('Average Order Value (Rs.)', avg_order)

    # ── Payment method breakdown ────────────────────────────────────────────
    s2_section('PAYMENT METHOD BREAKDOWN')
    s2_hdr('Method', 'Transactions', 'Amount (Rs.)')

    pay_summary = (
        Payment.objects
        .filter(paid_at__gte=start, paid_at__lte=end)
        .values('method')
        .annotate(cnt=Count('id'), tot=Sum('amount'))
        .order_by('-tot')
    )
    for i, row in enumerate(pay_summary):
        s2_row(row['method'], row['cnt'], float(row['tot'] or 0), alt=i % 2 == 1)

    cr_agg = (
        CreditRecord.objects
        .filter(record_type='CREDIT', order__created_at__gte=start,
                order__created_at__lte=end)
        .aggregate(cnt=Count('id'), tot=Sum('amount'))
    )
    if cr_agg['cnt']:
        s2_row('CREDIT', cr_agg['cnt'], float(cr_agg['tot'] or 0),
               alt=pay_summary.count() % 2 == 1)

    # ── Refunds ─────────────────────────────────────────────────────────────
    refunds_qs = (
        Refund.objects
        .filter(refunded_at__gte=start, refunded_at__lte=end)
        .select_related('payment__order')
        .order_by('refunded_at')
    )
    if refunds_qs.exists():
        s2_section('REFUNDS IN PERIOD')
        s2_hdr('Date', 'Order No', 'Method', 'Reason', 'Amount (Rs.)')
        for i, ref in enumerate(refunds_qs):
            lt = timezone.localtime(ref.refunded_at)
            ws2.cell(row=r, column=1, value=lt.strftime('%d %b %Y %H:%M')).font = data_font()
            ws2.cell(row=r, column=2, value=ref.payment.order.order_no).font = data_font()
            ws2.cell(row=r, column=3, value=ref.payment.method).font = data_font()
            ws2.cell(row=r, column=4, value=ref.reason or '—').font = data_font()
            ac = ws2.cell(row=r, column=5, value=float(ref.amount))
            ac.font = data_font(); ac.number_format = MONEY; ac.alignment = RIGHT
            if i % 2 == 1:
                for c in range(1, 6):
                    ws2.cell(row=r, column=c).fill = ALT_FILL
            ws2.row_dimensions[r].height = 18
            r += 1

    # ── Daily breakdown ─────────────────────────────────────────────────────
    s2_section('DAILY BREAKDOWN')
    s2_hdr('Date', 'Day', 'Orders', 'Gross (Rs.)', 'Discount (Rs.)', 'Net (Rs.)')

    tz_info  = timezone.get_current_timezone()
    cur_date = date_from
    day_idx  = 0
    while cur_date <= date_to:
        ds = timezone.datetime.combine(cur_date, timezone.datetime.min.time()).replace(tzinfo=tz_info)
        de = timezone.datetime.combine(cur_date, timezone.datetime.max.time()).replace(tzinfo=tz_info)
        day_orders = [o for o in orders
                      if ds <= o.created_at <= de]
        day_gross  = sum(o.grand_total for o in day_orders)
        day_disc   = sum(o.discount_amount for o in day_orders)
        day_net    = day_gross - day_disc
        s2_row(
            cur_date.strftime('%d %b %Y'),
            cur_date.strftime('%A'),
            len(day_orders),
            float(day_gross),
            float(day_disc),
            float(day_net),
            alt=day_idx % 2 == 1,
        )
        cur_date += timedelta(days=1)
        day_idx  += 1

    # Grand total row for daily section
    ws2.cell(row=r, column=1, value='TOTAL').font = tot_font()
    ws2.cell(row=r, column=1).fill = TOT_FILL
    ws2.cell(row=r, column=1).alignment = CENTER
    ws2.merge_cells(f'A{r}:C{r}')
    for c, val in [(4, float(tot_grand)), (5, float(tot_disc)), (6, float(tot_net))]:
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = tot_font(); cell.fill = TOT_FILL
        cell.number_format = MONEY; cell.alignment = RIGHT
    ws2.row_dimensions[r].height = 22

    # ── Serve the file ───────────────────────────────────────────────────────
    fname = f'Yasumi_Master_Report_{date_from}_{date_to}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
